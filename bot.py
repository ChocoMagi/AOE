import os
import re
import sqlite3
import asyncio

import discord
from discord import app_commands

# =========================================================
# CONFIG
# =========================================================

DB_PATH = "silver.db"
BACKUP_XLSX_PATH = os.getenv("SILVER_BACKUP_XLSX", "backups/silver_backup_latest.xlsx")
_OPENPYXL_MISSING_WARNED = False
USER_ID_COLUMN_NAMES = {"user_id", "initiator_id", "sender_id", "receiver_id", "recipient_id"}
GUILD_ID_COLUMN_NAMES = {"guild_id"}

# =========================================================
# ENV LOADING
# =========================================================

def load_env(path: str = ".env"):
    """Load key=value pairs from a .env file without overriding existing env."""
    if not os.path.exists(path):
        return

    with open(path, "r", encoding="utf-8") as file:
        for line in file:
            raw = line.strip()
            if not raw or raw.startswith("#") or "=" not in raw:
                continue

            key, value = raw.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_env()
TOKEN = os.getenv("DISCORD_TOKEN")
DEV_GUILD_ID = os.getenv("DEV_GUILD_ID")
GUILD_OBJECT = discord.Object(id=int(DEV_GUILD_ID)) if DEV_GUILD_ID else None

# =========================================================
# DATABASE SETUP
# =========================================================

class Database:
    def __init__(self, path: str):
        self.path = path

    def connect(self) -> sqlite3.Connection:
        return sqlite3.connect(self.path)

    @staticmethod
    def _ensure_account(conn: sqlite3.Connection, guild_id: int, user_id: int):
        conn.execute("""
            INSERT OR IGNORE INTO accounts (guild_id, user_id, wallet)
            VALUES (?, ?, 0)
        """, (guild_id, user_id))

    @staticmethod
    def _ensure_treasury(conn: sqlite3.Connection, guild_id: int):
        conn.execute("""
            INSERT OR IGNORE INTO treasury (guild_id, balance)
            VALUES (?, 0)
        """, (guild_id,))

    @staticmethod
    def _get_wallet(conn: sqlite3.Connection, guild_id: int, user_id: int) -> int:
        row = conn.execute(
            "SELECT wallet FROM accounts WHERE guild_id = ? AND user_id = ?",
            (guild_id, user_id),
        ).fetchone()
        return row[0] if row else 0

    def init_schema(self):
        """Create core tables if they do not exist."""
        with self.connect() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS accounts (
                    guild_id INTEGER NOT NULL,
                    user_id INTEGER NOT NULL,
                    wallet INTEGER NOT NULL DEFAULT 0,
                    PRIMARY KEY (guild_id, user_id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS lootsplit_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    guild_id INTEGER NOT NULL,
                    initiator_id INTEGER NOT NULL,
                    lootsplit_name TEXT,
                    total INTEGER NOT NULL,
                    tax_percent INTEGER NOT NULL,
                    tax_amount INTEGER NOT NULL,
                    remaining INTEGER NOT NULL,
                    share INTEGER NOT NULL,
                    recipient_count INTEGER NOT NULL,
                    recipient_ids TEXT NOT NULL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            """)
            columns = {
                row[1]
                for row in conn.execute("PRAGMA table_info(lootsplit_logs)").fetchall()
            }
            if "lootsplit_name" not in columns:
                conn.execute("ALTER TABLE lootsplit_logs ADD COLUMN lootsplit_name TEXT")
            conn.execute("""
                CREATE TABLE IF NOT EXISTS lootsplit_recipients (
                    lootsplit_id INTEGER NOT NULL,
                    recipient_id INTEGER NOT NULL,
                    PRIMARY KEY (lootsplit_id, recipient_id)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS transfer_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    guild_id INTEGER NOT NULL,
                    sender_id INTEGER NOT NULL,
                    receiver_id INTEGER NOT NULL,
                    amount INTEGER NOT NULL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS treasury_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    guild_id INTEGER NOT NULL,
                    initiator_id INTEGER NOT NULL,
                    action TEXT NOT NULL,
                    amount INTEGER NOT NULL,
                    recipient_id INTEGER,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS treasury (
                    guild_id INTEGER PRIMARY KEY,
                    balance INTEGER NOT NULL DEFAULT 0
                )
            """)

    def ensure_account(self, guild_id: int, user_id: int):
        """Ensure a user has an account row, creating it with 0 silver if missing."""
        with self.connect() as conn:
            self._ensure_account(conn, guild_id, user_id)

    def get_balance(self, guild_id: int, user_id: int) -> int:
        """Read and return the user's wallet balance."""
        with self.connect() as conn:
            return self._get_wallet(conn, guild_id, user_id)

    def add_treasury(self, guild_id: int, amount: int):
        """Add to the guild treasury balance."""
        with self.connect() as conn:
            conn.execute("""
                INSERT INTO treasury (guild_id, balance)
                VALUES (?, ?)
                ON CONFLICT(guild_id) DO UPDATE SET balance = balance + excluded.balance
            """, (guild_id, amount))

    def deduct_treasury(self, guild_id: int, amount: int) -> bool:
        """Subtract from the guild treasury; return False if it would go negative."""
        with self.connect() as conn:
            try:
                conn.execute("BEGIN")
                self._ensure_treasury(conn, guild_id)
                row = conn.execute(
                    "SELECT balance FROM treasury WHERE guild_id = ?",
                    (guild_id,),
                ).fetchone()
                current = row[0] if row else 0
                if current < amount:
                    conn.execute("ROLLBACK")
                    return False
                conn.execute(
                    "UPDATE treasury SET balance = balance - ? WHERE guild_id = ?",
                    (amount, guild_id),
                )
                conn.execute("COMMIT")
                return True
            except sqlite3.Error:
                conn.execute("ROLLBACK")
                raise

    def transfer_treasury_to_user(self, guild_id: int, user_id: int, amount: int) -> bool:
        """Move silver from treasury to a user; return False if treasury lacks funds."""
        with self.connect() as conn:
            try:
                conn.execute("BEGIN")
                self._ensure_treasury(conn, guild_id)
                self._ensure_account(conn, guild_id, user_id)
                row = conn.execute(
                    "SELECT balance FROM treasury WHERE guild_id = ?",
                    (guild_id,),
                ).fetchone()
                current = row[0] if row else 0
                if current < amount:
                    conn.execute("ROLLBACK")
                    return False
                conn.execute(
                    "UPDATE treasury SET balance = balance - ? WHERE guild_id = ?",
                    (amount, guild_id),
                )
                conn.execute(
                    "UPDATE accounts SET wallet = wallet + ? WHERE guild_id = ? AND user_id = ?",
                    (amount, guild_id, user_id),
                )
                conn.execute("COMMIT")
                return True
            except sqlite3.Error:
                conn.execute("ROLLBACK")
                raise

    def get_treasury(self, guild_id: int) -> int:
        """Read and return the guild treasury balance."""
        with self.connect() as conn:
            row = conn.execute(
                "SELECT balance FROM treasury WHERE guild_id = ?",
                (guild_id,),
            ).fetchone()
            return row[0] if row else 0

    def log_treasury(
        self,
        guild_id: int,
        initiator_id: int,
        action: str,
        amount: int,
        recipient_id: int | None = None,
    ):
        with self.connect() as conn:
            conn.execute("""
                INSERT INTO treasury_logs (guild_id, initiator_id, action, amount, recipient_id)
                VALUES (?, ?, ?, ?, ?)
            """, (guild_id, initiator_id, action, amount, recipient_id))

    def get_treasury_history(self, guild_id: int, limit: int | None, offset: int = 0):
        with self.connect() as conn:
            if limit is None:
                return conn.execute("""
                    SELECT initiator_id, action, amount, recipient_id, created_at
                    FROM treasury_logs
                    WHERE guild_id = ?
                    ORDER BY id DESC
                """, (guild_id,)).fetchall()
            return conn.execute("""
                SELECT initiator_id, action, amount, recipient_id, created_at
                FROM treasury_logs
                WHERE guild_id = ?
                ORDER BY id DESC
                LIMIT ? OFFSET ?
            """, (guild_id, limit, offset)).fetchall()

    def add_balance(self, guild_id: int, user_id: int, amount: int) -> bool:
        """Add (or subtract) silver; return False if it would go negative."""
        with self.connect() as conn:
            current = self._get_wallet(conn, guild_id, user_id)
            new_balance = current + amount
            if new_balance < 0:
                return False
            conn.execute(
                "UPDATE accounts SET wallet = ? WHERE guild_id = ? AND user_id = ?",
                (new_balance, guild_id, user_id),
            )
            return True

    def transfer_balance(self, guild_id: int, sender_id: int, receiver_id: int, amount: int) -> bool:
        """Atomically move silver; return False if sender lacks funds."""
        with self.connect() as conn:
            try:
                conn.execute("BEGIN")
                self._ensure_account(conn, guild_id, sender_id)
                self._ensure_account(conn, guild_id, receiver_id)

                current = self._get_wallet(conn, guild_id, sender_id)
                if current < amount:
                    conn.execute("ROLLBACK")
                    return False

                conn.execute(
                    "UPDATE accounts SET wallet = wallet - ? WHERE guild_id = ? AND user_id = ?",
                    (amount, guild_id, sender_id),
                )
                conn.execute(
                    "UPDATE accounts SET wallet = wallet + ? WHERE guild_id = ? AND user_id = ?",
                    (amount, guild_id, receiver_id),
                )
                conn.execute("COMMIT")
                return True
            except sqlite3.Error:
                conn.execute("ROLLBACK")
                raise

    def deduct_balance(self, guild_id: int, user_id: int, amount: int) -> bool:
        """Atomically subtract silver; return False if it would go negative."""
        with self.connect() as conn:
            try:
                conn.execute("BEGIN")
                self._ensure_account(conn, guild_id, user_id)

                current = self._get_wallet(conn, guild_id, user_id)
                if current < amount:
                    conn.execute("ROLLBACK")
                    return False

                conn.execute(
                    "UPDATE accounts SET wallet = wallet - ? WHERE guild_id = ? AND user_id = ?",
                    (amount, guild_id, user_id),
                )
                conn.execute("COMMIT")
                return True
            except sqlite3.Error:
                conn.execute("ROLLBACK")
                raise

    def log_lootsplit(
        self,
        guild_id: int,
        initiator_id: int,
        lootsplit_name: str | None,
        total: int,
        tax_percent: int,
        tax_amount: int,
        remaining: int,
        share: int,
        recipient_count: int,
        recipient_ids: str,
    ):
        with self.connect() as conn:
            cursor = conn.execute("""
                INSERT INTO lootsplit_logs (
                    guild_id,
                    initiator_id,
                    lootsplit_name,
                    total,
                    tax_percent,
                    tax_amount,
                    remaining,
                    share,
                    recipient_count,
                    recipient_ids
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                guild_id,
                initiator_id,
                lootsplit_name,
                total,
                tax_percent,
                tax_amount,
                remaining,
                share,
                recipient_count,
                recipient_ids,
            ))
            lootsplit_id = cursor.lastrowid
            if lootsplit_id is not None and recipient_ids:
                rows = [(lootsplit_id, int(uid)) for uid in recipient_ids.split(",") if uid]
                conn.executemany(
                    "INSERT OR IGNORE INTO lootsplit_recipients (lootsplit_id, recipient_id) VALUES (?, ?)",
                    rows,
                )

    def log_transfer(self, guild_id: int, sender_id: int, receiver_id: int, amount: int):
        with self.connect() as conn:
            conn.execute("""
                INSERT INTO transfer_logs (guild_id, sender_id, receiver_id, amount)
                VALUES (?, ?, ?, ?)
            """, (guild_id, sender_id, receiver_id, amount))

    def get_lootsplit_history(self, guild_id: int, limit: int, offset: int):
        with self.connect() as conn:
            return conn.execute("""
                SELECT
                    l.initiator_id,
                    l.lootsplit_name,
                    l.total,
                    l.tax_percent,
                    l.share,
                    COALESCE(GROUP_CONCAT(r.recipient_id), l.recipient_ids) AS recipient_ids,
                    l.created_at
                FROM lootsplit_logs l
                LEFT JOIN lootsplit_recipients r
                  ON r.lootsplit_id = l.id
                WHERE l.guild_id = ?
                GROUP BY l.id
                ORDER BY l.id DESC
                LIMIT ? OFFSET ?
            """, (guild_id, limit, offset)).fetchall()

    def get_transfer_history(self, guild_id: int, limit: int, offset: int):
        with self.connect() as conn:
            return conn.execute("""
                SELECT sender_id, receiver_id, amount, created_at
                FROM transfer_logs
                WHERE guild_id = ?
                ORDER BY id DESC
                LIMIT ? OFFSET ?
            """, (guild_id, limit, offset)).fetchall()

    def get_leaderboard(self, guild_id: int, limit: int, offset: int):
        with self.connect() as conn:
            return conn.execute("""
                SELECT user_id, wallet
                FROM accounts
                WHERE guild_id = ?
                  AND wallet > 0
                ORDER BY wallet DESC, user_id ASC
                LIMIT ? OFFSET ?
            """, (guild_id, limit, offset)).fetchall()

    def get_total_silver(self, guild_id: int) -> int:
        with self.connect() as conn:
            row = conn.execute("""
                SELECT COALESCE(SUM(wallet), 0)
                FROM accounts
                WHERE guild_id = ?
            """, (guild_id,)).fetchone()
            return row[0] if row else 0

    def get_leaderboard_count(self, guild_id: int) -> int:
        with self.connect() as conn:
            row = conn.execute("""
                SELECT COUNT(*)
                FROM accounts
                WHERE guild_id = ?
                  AND wallet > 0
            """, (guild_id,)).fetchone()
            return row[0] if row else 0


db = Database(DB_PATH)


def clamp_limit(limit: int, max_limit: int = 10) -> int:
    return max(1, min(max_limit, limit))


def clamp_page(page: int) -> int:
    return max(1, page)


async def send_error(interaction: discord.Interaction, message: str):
    return await interaction.response.send_message(message, ephemeral=True)


def format_silver(amount: int) -> str:
    return f"{amount:,}"


def _build_lootsplit_lines(
    total: int,
    repair: int,
    tax: int,
    tax_amount: int,
    remaining: int,
    share: int,
    recipients: list[discord.Member],
    lootsplit_name: str | None,
) -> list[str]:
    lines: list[str] = []
    if lootsplit_name:
        lines.append(f"Lootsplit name: **{lootsplit_name}**")
    lines.extend(
        [
            f"Total: **{format_silver(total)} silver**",
            f"Repair: **{format_silver(repair)} silver**",
            f"Tax ({tax}%): **{format_silver(tax_amount)} silver**",
            f"Split: **{format_silver(remaining)} silver** among {', '.join(m.mention for m in recipients)}",
            f"Each receives **{format_silver(share)} silver**.",
        ]
    )
    return lines


class LootsplitConfirmView(discord.ui.View):
    def __init__(
        self,
        guild_id: int,
        guild: discord.Guild | None,
        initiator_id: int,
        lootsplit_name: str | None,
        total: int,
        repair: int,
        tax: int,
        tax_amount: int,
        remaining: int,
        share: int,
        recipients: list[discord.Member],
    ):
        super().__init__(timeout=120)
        self.guild_id = guild_id
        self.guild = guild
        self.initiator_id = initiator_id
        self.lootsplit_name = lootsplit_name
        self.total = total
        self.repair = repair
        self.tax = tax
        self.tax_amount = tax_amount
        self.remaining = remaining
        self.share = share
        self.recipients = recipients
        self.applied = False

    def _summary(self) -> str:
        return "\n".join(
            _build_lootsplit_lines(
                total=self.total,
                repair=self.repair,
                tax=self.tax,
                tax_amount=self.tax_amount,
                remaining=self.remaining,
                share=self.share,
                recipients=self.recipients,
                lootsplit_name=self.lootsplit_name,
            )
        )

    async def _reject_non_owner(self, interaction: discord.Interaction) -> bool:
        if interaction.user.id == self.initiator_id:
            return False
        await interaction.response.send_message(
            "Only the command initiator can confirm or cancel this lootsplit.",
            ephemeral=True,
        )
        return True

    def _disable_all_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                child.disabled = True

    @discord.ui.button(label="Confirm", style=discord.ButtonStyle.success)
    async def confirm(self, interaction: discord.Interaction, _: discord.ui.Button):
        if await self._reject_non_owner(interaction):
            return
        if self.applied:
            await interaction.response.send_message("This lootsplit was already confirmed.", ephemeral=True)
            return

        recipient_ids = ",".join(str(member.id) for member in self.recipients)
        for member in self.recipients:
            db.ensure_account(self.guild_id, member.id)
            db.add_balance(self.guild_id, member.id, self.share)

        if self.tax_amount > 0:
            db.add_treasury(self.guild_id, self.tax_amount)

        db.log_lootsplit(
            self.guild_id,
            self.initiator_id,
            self.lootsplit_name,
            self.total,
            self.tax,
            self.tax_amount,
            self.remaining,
            self.share,
            len(self.recipients),
            recipient_ids,
        )
        await run_excel_backup(self.guild)

        self.applied = True
        self._disable_all_buttons()
        await interaction.response.edit_message(
            content="Lootsplit confirmed and applied.\n" + self._summary(),
            view=self,
            allowed_mentions=discord.AllowedMentions.none(),
        )

    @discord.ui.button(label="Cancel", style=discord.ButtonStyle.danger)
    async def cancel(self, interaction: discord.Interaction, _: discord.ui.Button):
        if await self._reject_non_owner(interaction):
            return
        if self.applied:
            await interaction.response.send_message("This lootsplit was already confirmed.", ephemeral=True)
            return

        self._disable_all_buttons()
        await interaction.response.edit_message(
            content="Lootsplit canceled. No balances were changed.",
            view=self,
        )


def _quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _format_user_id(value: object, user_name_map: dict[int, str]) -> object:
    if not isinstance(value, int):
        return value
    name = user_name_map.get(value)
    if not name:
        return value
    return f"{name} ({value})"


def _format_recipient_ids(value: object, user_name_map: dict[int, str]) -> object:
    if not isinstance(value, str):
        return value
    parts = [p.strip() for p in value.split(",") if p.strip()]
    if not parts:
        return value

    formatted: list[str] = []
    for part in parts:
        if not part.isdigit():
            formatted.append(part)
            continue
        uid = int(part)
        name = user_name_map.get(uid)
        formatted.append(f"{name} ({uid})" if name else part)
    return ", ".join(formatted)


def _collect_user_ids_for_backup(db_path: str = DB_PATH) -> set[int]:
    user_ids: set[int] = set()
    try:
        with sqlite3.connect(db_path) as conn:
            queries = [
                ("SELECT user_id FROM accounts", False),
                ("SELECT initiator_id, recipient_ids FROM lootsplit_logs", True),
                ("SELECT recipient_id FROM lootsplit_recipients", False),
                ("SELECT sender_id, receiver_id FROM transfer_logs", False),
                ("SELECT initiator_id, recipient_id FROM treasury_logs", False),
            ]
            for query, has_recipient_ids in queries:
                for row in conn.execute(query).fetchall():
                    for value in row:
                        if isinstance(value, int):
                            user_ids.add(value)
                    if has_recipient_ids and len(row) > 1 and isinstance(row[1], str):
                        for part in row[1].split(","):
                            part = part.strip()
                            if part.isdigit():
                                user_ids.add(int(part))
    except Exception:
        return set()
    return user_ids


def export_database_to_excel(
    db_path: str = DB_PATH,
    output_path: str = BACKUP_XLSX_PATH,
    user_name_map: dict[int, str] | None = None,
    guild_name_map: dict[int, str] | None = None,
) -> tuple[bool, str]:
    user_name_map = user_name_map or {}
    guild_name_map = guild_name_map or {}
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError:
        return False, "openpyxl is not installed"

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    try:
        with sqlite3.connect(db_path) as conn:
            table_rows = conn.execute("""
                SELECT name
                FROM sqlite_master
                WHERE type = 'table'
                  AND name NOT LIKE 'sqlite_%'
                ORDER BY name
            """).fetchall()

            table_names = [row[0] for row in table_rows]

            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            if not table_names:
                sheet = workbook.create_sheet("empty")
                sheet.append(["message"])
                sheet.append(["No tables found in database"])
            else:
                for table_name in table_names:
                    query_table = _quote_identifier(table_name)
                    sheet_name = table_name[:31] or "sheet"
                    sheet = workbook.create_sheet(sheet_name)

                    header_cursor = conn.execute(f"SELECT * FROM {query_table} LIMIT 0")
                    headers = [col[0] for col in (header_cursor.description or [])]
                    if headers:
                        sheet.append(headers)

                    row_cursor = conn.execute(f"SELECT * FROM {query_table}")
                    for row in row_cursor.fetchall():
                        out_row = []
                        for index, value in enumerate(row):
                            column = headers[index] if index < len(headers) else ""
                            if column in USER_ID_COLUMN_NAMES:
                                out_row.append(_format_user_id(value, user_name_map))
                            elif column == "recipient_ids":
                                out_row.append(_format_recipient_ids(value, user_name_map))
                            elif column in GUILD_ID_COLUMN_NAMES and isinstance(value, int):
                                guild_name = guild_name_map.get(value)
                                out_row.append(f"{guild_name} ({value})" if guild_name else value)
                            else:
                                out_row.append(value)
                        sheet.append(out_row)

            workbook.save(output_path)
    except Exception as exc:
        return False, str(exc)

    return True, output_path


async def run_excel_backup(guild: discord.Guild | None = None) -> None:
    global _OPENPYXL_MISSING_WARNED

    user_name_map: dict[int, str] = {}
    guild_name_map: dict[int, str] = {}

    if guild is not None:
        guild_name_map[guild.id] = guild.name
        user_ids = await asyncio.to_thread(_collect_user_ids_for_backup)
        for user_id in user_ids:
            member = guild.get_member(user_id)
            if member is None:
                try:
                    member = await guild.fetch_member(user_id)
                except (discord.NotFound, discord.Forbidden, discord.HTTPException):
                    member = None
            if member is not None:
                user_name_map[user_id] = str(member)

    success, detail = await asyncio.to_thread(
        export_database_to_excel,
        DB_PATH,
        BACKUP_XLSX_PATH,
        user_name_map,
        guild_name_map,
    )
    if success:
        return

    if detail == "openpyxl is not installed":
        if not _OPENPYXL_MISSING_WARNED:
            print("Excel backup disabled: install openpyxl (`pip install openpyxl`).")
            _OPENPYXL_MISSING_WARNED = True
        return

    print(f"Excel backup failed: {detail}")



# =========================================================
# BOT SETUP
# =========================================================

intents = discord.Intents.default()
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)

# =========================================================
# BOT EVENTS
# =========================================================

@tree.error
async def on_app_command_error(
    interaction: discord.Interaction,
    error: app_commands.AppCommandError,
):
    if isinstance(error, app_commands.MissingPermissions):
        return await send_error(
            interaction,
            "You need administrator privileges to use this command.",
        )
    raise error

@client.event
async def on_ready():
    db.init_schema()
    tree.clear_commands(guild=None)
    await tree.sync()
    synced = await tree.sync(guild=GUILD_OBJECT)
    print(f"Synced {len(synced)} commands to dev guild | Logged in as {client.user}")

# =========================================================

if not TOKEN:
    raise RuntimeError("DISCORD_TOKEN is not set")
if not DEV_GUILD_ID:
    raise RuntimeError("DEV_GUILD_ID is not set")

@tree.command(guild=GUILD_OBJECT, name="balance", description="Check a user's silver balance")
@app_commands.describe(member="User to check (defaults to you)")
async def balance(interaction: discord.Interaction, member: discord.Member | None = None):
    """Show a user's silver (defaults to the caller)."""
    target = member or interaction.user
    db.ensure_account(interaction.guild.id, target.id)
    wallet = db.get_balance(interaction.guild.id, target.id)
    if target.id == interaction.user.id:
        await interaction.response.send_message(f"You have **{format_silver(wallet)} silver**")
    else:
        await interaction.response.send_message(
            f"{target.mention} has **{format_silver(wallet)} silver**",
            allowed_mentions=discord.AllowedMentions.none(),
        )

@tree.command(guild=GUILD_OBJECT, name="treasury", description="Check the guild treasury balance")
@app_commands.checks.has_permissions(administrator=True)
async def treasury(interaction: discord.Interaction):
    """Show the guild treasury balance."""
    if interaction.guild is None:
        return await send_error(interaction, "This command can only be used in a server.")

    balance_amount = db.get_treasury(interaction.guild.id)
    await interaction.response.send_message(
        f"Guild treasury has **{format_silver(balance_amount)} silver**."
    )

@tree.command(guild=GUILD_OBJECT, name="purge", description="Delete recent messages in this channel")
@app_commands.describe(limit="Number of messages to delete (max 100)")
@app_commands.checks.has_permissions(manage_messages=True)
async def purge_messages(interaction: discord.Interaction, limit: int):
    """Delete recent messages in the current channel (manage messages only)."""
    if interaction.guild is None:
        return await send_error(interaction, "This command can only be used in a server.")

    channel = interaction.channel
    if not isinstance(channel, discord.TextChannel):
        return await send_error(interaction, "This command can only be used in a text channel.")

    limit = clamp_limit(limit, 100)
    await interaction.response.defer(ephemeral=True)
    deleted = await channel.purge(limit=limit)
    await interaction.followup.send(
        f"Deleted **{len(deleted)}** messages in {channel.mention}.",
        ephemeral=True,
    )

@tree.command(guild=GUILD_OBJECT, name="lootsplit", description="Split silver among mentioned users after repair fee and tax")
@app_commands.describe(
    total="Total silver",
    repair="Flat repair fee",
    tax="Tax percent",
    users="Mention users",
    name="Optional name for this lootsplit",
)
@app_commands.checks.has_permissions(administrator=True)
async def lootsplit(
    interaction: discord.Interaction,
    total: int,
    repair: int,
    tax: int,
    users: str,
    name: str | None = None,
):
    """Split a total amount after a flat repair fee and tax, then distribute evenly."""
    if interaction.guild is None:
        return await send_error(interaction, "This command can only be used in a server.")

    if total <= 0:
        return await send_error(interaction, "Total must be positive.")

    if repair < 0:
        return await send_error(interaction, "Repair must be 0 or higher.")

    if not 0 <= tax <= 100:
        return await send_error(interaction, "Tax must be 0-100.")

    lootsplit_name = (name or "").strip() or None
    if lootsplit_name and len(lootsplit_name) > 80:
        return await send_error(interaction, "Lootsplit name must be 80 characters or fewer.")

    user_ids = set(int(uid) for uid in re.findall(r"<@!?(\d+)>", users))
    if not user_ids:
        user_ids = set(int(uid) for uid in re.findall(r"\b(\d{17,20})\b", users))

    recipients = []
    if not user_ids:
        for token in users.split():
            if not token.startswith("@"):
                continue
            member = interaction.guild.get_member_named(token[1:])
            if member is not None:
                recipients.append(member)

    forbidden_count = 0
    not_found_count = 0
    http_error_count = 0
    for uid in sorted(user_ids):
        member = interaction.guild.get_member(uid)
        if member is None:
            try:
                member = await interaction.guild.fetch_member(uid)
            except discord.NotFound:
                not_found_count += 1
                continue
            except discord.Forbidden:
                forbidden_count += 1
                continue
            except discord.HTTPException:
                http_error_count += 1
                continue
        recipients.append(member)

    if not recipients:
        if user_ids:
            preview_ids = ", ".join(str(uid) for uid in list(user_ids)[:5])
            return await send_error(
                interaction,
                "No valid users found. Make sure the users are in this server. "
                f"IDs: {preview_ids} | NotFound: {not_found_count} "
                f"Forbidden: {forbidden_count} HTTP: {http_error_count}",
            )
        if not user_ids:
            return await send_error(
                interaction,
                "Provide mentions or user IDs. Example: @User1 @User2 or 123... 456...",
            )
        return await send_error(
            interaction,
            "No valid users found. Make sure the users are in this server.",
        )

    after_repair = total - repair
    if after_repair <= 0:
        return await send_error(interaction, "Repair fee is too high for this total.")

    tax_amount = (after_repair * tax) // 100
    remaining = after_repair - tax_amount
    share = remaining // len(recipients)

    if share <= 0:
        return await send_error(interaction, "Not enough silver to split.")

    preview_lines = _build_lootsplit_lines(
        total=total,
        repair=repair,
        tax=tax,
        tax_amount=tax_amount,
        remaining=remaining,
        share=share,
        recipients=recipients,
        lootsplit_name=lootsplit_name,
    )
    preview_lines.append("Press **Confirm** to apply this lootsplit, or **Cancel** to abort.")
    view = LootsplitConfirmView(
        interaction.guild.id,
        interaction.guild,
        interaction.user.id,
        lootsplit_name,
        total,
        repair,
        tax,
        tax_amount,
        remaining,
        share,
        recipients,
    )
    await interaction.response.send_message(
        "\n".join(preview_lines),
        view=view,
        allowed_mentions=discord.AllowedMentions.none(),
    )

@tree.command(guild=GUILD_OBJECT, name="give_silver", description="Add silver to a user's balance")
@app_commands.describe(member="User to receive silver", amount="Amount to add")
@app_commands.checks.has_permissions(administrator=True)
async def give_silver(interaction: discord.Interaction, member: discord.Member, amount: int):
    """Manually add silver to a user's wallet (admin only)."""
    if amount <= 0:
        return await send_error(interaction, "Amount must be positive.")

    db.ensure_account(interaction.guild.id, member.id)
    db.add_balance(interaction.guild.id, member.id, amount)
    await run_excel_backup(interaction.guild)

    await interaction.response.send_message(
        f"Added **{format_silver(amount)} silver** to {member.mention}."
    )

@tree.command(guild=GUILD_OBJECT, name="treasury_add", description="Add silver to the treasury")
@app_commands.describe(amount="Amount to add")
@app_commands.checks.has_permissions(administrator=True)
async def treasury_add(interaction: discord.Interaction, amount: int):
    """Add silver to the guild treasury (admin only)."""
    if amount <= 0:
        return await send_error(interaction, "Amount must be positive.")

    db.add_treasury(interaction.guild.id, amount)
    db.log_treasury(interaction.guild.id, interaction.user.id, "add", amount)
    await run_excel_backup(interaction.guild)
    await interaction.response.send_message(
        f"Added **{format_silver(amount)} silver** to the treasury."
    )

@tree.command(guild=GUILD_OBJECT, name="take_silver", description="Remove silver from a user's balance")
@app_commands.describe(member="User to remove silver from", amount="Amount to remove")
@app_commands.checks.has_permissions(administrator=True)
async def take_silver(interaction: discord.Interaction, member: discord.Member, amount: int):
    """Manually remove silver from a user's wallet (admin only)."""
    if amount <= 0:
        return await send_error(interaction, "Amount must be positive.")

    if not db.deduct_balance(interaction.guild.id, member.id, amount):
        return await send_error(interaction, "Insufficient balance; no silver was removed.")
    await run_excel_backup(interaction.guild)

    await interaction.response.send_message(
        f"Removed **{format_silver(amount)} silver** from {member.mention}."
    )

@tree.command(guild=GUILD_OBJECT, name="treasury_take", description="Remove silver from the treasury")
@app_commands.describe(amount="Amount to remove", member="Optional user to receive the silver")
@app_commands.checks.has_permissions(administrator=True)
async def treasury_take(
    interaction: discord.Interaction,
    amount: int,
    member: discord.Member | None = None,
):
    """Remove silver from the treasury (admin only)."""
    if amount <= 0:
        return await send_error(interaction, "Amount must be positive.")

    if member is not None:
        success = db.transfer_treasury_to_user(interaction.guild.id, member.id, amount)
        if not success:
            return await send_error(interaction, "Treasury has insufficient funds.")
        db.log_treasury(interaction.guild.id, interaction.user.id, "transfer", amount, member.id)
        await run_excel_backup(interaction.guild)
        await interaction.response.send_message(
            f"Transferred **{format_silver(amount)} silver** from the treasury to {member.mention}."
        )
        return

    if not db.deduct_treasury(interaction.guild.id, amount):
        return await send_error(interaction, "Treasury has insufficient funds.")

    db.log_treasury(interaction.guild.id, interaction.user.id, "take", amount)
    await run_excel_backup(interaction.guild)
    await interaction.response.send_message(
        f"Removed **{format_silver(amount)} silver** from the treasury."
    )

@tree.command(guild=GUILD_OBJECT, name="transfer", description="Send silver to another user")
@app_commands.describe(member="User to receive silver", amount="Amount to send")
async def transfer(interaction: discord.Interaction, member: discord.Member, amount: int):
    """Move silver from the caller to another member."""
    if amount <= 0:
        return await send_error(interaction, "Amount must be positive.")

    if member.id == interaction.user.id:
        return await send_error(interaction, "You can't pay yourself.")

    if not db.transfer_balance(interaction.guild.id, interaction.user.id, member.id, amount):
        return await send_error(interaction, "Not enough silver.")

    db.log_transfer(interaction.guild.id, interaction.user.id, member.id, amount)
    await run_excel_backup(interaction.guild)

    await interaction.response.send_message(
        f"{interaction.user.mention} sent **{format_silver(amount)} silver** to {member.mention}."
    )

@tree.command(guild=GUILD_OBJECT, name="lootsplit_history", description="Show recent lootsplit history")
@app_commands.describe(limit="Number of entries per page (max 10)", page="Page number")
@app_commands.checks.has_permissions(administrator=True)
async def lootsplit_history(interaction: discord.Interaction, limit: int = 5, page: int = 1):
    """Show recent lootsplit logs for this guild."""
    limit = clamp_limit(limit)
    page = clamp_page(page)
    offset = (page - 1) * limit
    rows = db.get_lootsplit_history(interaction.guild.id, limit, offset)

    if not rows:
        return await interaction.response.send_message("No lootsplit history yet.")

    lines = [f"Recent lootsplits (page {page}):"]
    for initiator_id, lootsplit_name, total, tax_percent, share, recipient_ids, created_at in rows:
        recipients = ", ".join(f"<@{uid}>" for uid in recipient_ids.split(",") if uid)
        name_prefix = f"[{lootsplit_name}] " if lootsplit_name else ""
        lines.append(
            f"{created_at} - <@{initiator_id}> {name_prefix}split {format_silver(total)} silver "
            f"(tax {tax_percent}%), {recipients}, {format_silver(share)} each"
        )

    await interaction.response.send_message(
        "\n".join(lines),
        allowed_mentions=discord.AllowedMentions.none(),
    )


@tree.command(guild=GUILD_OBJECT, name="transfer_history", description="Show recent transfer history")
@app_commands.describe(limit="Number of entries per page (max 10)", page="Page number")
@app_commands.checks.has_permissions(administrator=True)
async def transfer_history(interaction: discord.Interaction, limit: int = 5, page: int = 1):
    """Show recent transfer logs for this guild."""
    limit = clamp_limit(limit)
    page = clamp_page(page)
    offset = (page - 1) * limit
    rows = db.get_transfer_history(interaction.guild.id, limit, offset)

    if not rows:
        return await interaction.response.send_message("No transfer history yet.")

    lines = [f"Recent transfers (page {page}):"]
    for sender_id, receiver_id, amount, created_at in rows:
        lines.append(
            f"{created_at} - <@{sender_id}> -> <@{receiver_id}> : {format_silver(amount)} silver"
        )

    await interaction.response.send_message(
        "\n".join(lines),
        allowed_mentions=discord.AllowedMentions.none(),
    )

@tree.command(guild=GUILD_OBJECT, name="treasury_history", description="Show recent treasury activity")
@app_commands.describe(limit="Optional number of entries per page (max 10)", page="Page number")
@app_commands.checks.has_permissions(administrator=True)
async def treasury_history(interaction: discord.Interaction, limit: int | None = None, page: int = 1):
    """Show recent treasury logs for this guild."""
    offset = 0
    if limit is not None:
        limit = clamp_limit(limit)
        page = clamp_page(page)
        offset = (page - 1) * limit
    rows = db.get_treasury_history(interaction.guild.id, limit, offset)

    if not rows:
        return await interaction.response.send_message("No treasury activity yet.")

    header = f"Recent treasury activity (page {page}):" if limit is not None else "Recent treasury activity:"
    lines = [header]
    for initiator_id, action, amount, recipient_id, created_at in rows:
        if action == "transfer" and recipient_id:
            detail = f"-> <@{recipient_id}>"
        else:
            detail = ""
        lines.append(
            f"{created_at} - <@{initiator_id}> {action} {format_silver(amount)} silver {detail}".rstrip()
        )

    await interaction.response.send_message(
        "\n".join(lines),
        allowed_mentions=discord.AllowedMentions.none(),
    )

@tree.command(guild=GUILD_OBJECT, name="leaderboard", description="Show richest users in the server")
@app_commands.describe(page="Page number (10 per page)")
async def leaderboard(interaction: discord.Interaction, page: int = 1):
    """Show the top balances in this guild."""
    if interaction.guild is None:
        return await send_error(interaction, "This command can only be used in a server.")

    page = clamp_page(page)
    limit = 10
    offset = (page - 1) * limit
    rows = db.get_leaderboard(interaction.guild.id, limit, offset)
    total_silver = db.get_total_silver(interaction.guild.id)
    total_rows = db.get_leaderboard_count(interaction.guild.id)
    total_pages = max(1, (total_rows + limit - 1) // limit)

    if not rows:
        return await interaction.response.send_message("No balances yet.")

    if page > total_pages:
        return await send_error(interaction, f"Page out of range. Max page is {total_pages}.")

    lines = [
        f"Leaderboard (page {page}/{total_pages}) - Total owed: {format_silver(total_silver)} silver"
    ]
    start_rank = offset + 1
    for i, (user_id, wallet) in enumerate(rows, start=start_rank):
        lines.append(f"{i}. <@{user_id}> — {format_silver(wallet)} silver")

    await interaction.response.send_message(
        "\n".join(lines),
        allowed_mentions=discord.AllowedMentions.none(),
    )

@tree.command(guild=GUILD_OBJECT, name="guild_balance", description="Calculate the guild's actual balance")
@app_commands.describe(amount="Total silver on hand")
@app_commands.checks.has_permissions(administrator=True)
async def guild_balance(interaction: discord.Interaction, amount: int):
    """Calculate actual balance after treasury and owed balances."""
    if interaction.guild is None:
        return await send_error(interaction, "This command can only be used in a server.")

    if amount < 0:
        return await send_error(interaction, "Amount must be 0 or positive.")

    treasury_amount = db.get_treasury(interaction.guild.id)
    total_owed = db.get_total_silver(interaction.guild.id)
    actual_balance = amount - treasury_amount - total_owed

    await interaction.response.send_message(
        f"Actual balance: **{format_silver(actual_balance)} silver**\n"
        f"Total on hand: **{format_silver(amount)} silver**\n"
        f"Treasury: **{format_silver(treasury_amount)} silver**\n"
        f"Total owed: **{format_silver(total_owed)} silver**"
    )

@tree.command(guild=GUILD_OBJECT, name="sync", description="Sync application commands (admin only)")
@app_commands.checks.has_permissions(administrator=True)
async def sync_commands(interaction: discord.Interaction):
    """Force a command sync for this guild."""
    if interaction.guild is None:
        return await send_error(interaction, "This command can only be used in a server.")

    await interaction.response.defer(ephemeral=True, thinking=True)
    tree.clear_commands(guild=None)
    await tree.sync()
    await tree.sync(guild=interaction.guild)
    await interaction.followup.send("Commands synced.", ephemeral=True)

client.run(TOKEN)
